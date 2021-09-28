from openpyxl import load_workbook, Workbook
import csv

class ReadSeries:
    def __init__(self, name="series.csv"):
        self.name = name
        if '.csv' in name:
            self.type = 'csv'
            csvFile = open(name)
            self.csv = csv.reader(csvFile)
        elif '.xls' in name:
            self.type = 'exel'
            self.sheet = load_workbook(name).active

    def cell(self, row, col):
        if self.type == 'exel':
            return self.sheet.cell(row, col).value

    def rowInArr(self, row):
        arr = []
        for i in range(1, self.len(row)+1):
            cell = self.cell(row, i)
            if cell == None:
                break
            arr += [self.cell(row, i)]
        return arr

    def dataInList(self):
        if self.type == 'csv':
            return self.csvInList()
        elif self.type == 'exel':
            return self.exelInList()

    def exelInList(self):
        data = []
        numberOfRows = self.rowsNum()
        for i in range(1, numberOfRows+1):
            data.append(self.rowInArr(i))
        return data
    
    def csvInList(self):
        data = []
        for row in self.csv:
            if row[0]:
                data.append([float(col) for col in row if col])
        return data

    def len(self, rowNum):
        length = len(self.sheet[rowNum])
        while (self.cell(rowNum, length) == None):
            length = length - 1
        return length
    
    def rowsNum(self):
        return len(tuple((self.sheet.rows)))

